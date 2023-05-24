import random
from openpyxl import load_workbook
import config

# 定义起始位置和范围
start_row = 11
start_col = 2
num_rows = 1000  # 需要生成的行数

# 加载工作簿和工作表
wb = load_workbook(config.EXCEL_PATH)  # 使用你的文件名替换'filename.xlsx'
ws = wb['TwoSteps_sequential']  # 或者使用工作表名称：wb['Sheet1']

# 遍历每一行
for i in range(start_row, start_row + num_rows):
    # 在每一行的指定位置生成随机数
    ws.cell(row=i, column=start_col, value=(i-10)*0.02)  # 到达率
    ws.cell(row=i, column=start_col + 1, value=0.8)  # 服务时间
    ws.cell(row=i, column=start_col + 2, value=1)  # 服务能力，这里假设它是1到10的整数
    
    # 在每一行的指定位置生成随机数
    ws.cell(row=i, column=start_col + 8, value=0.26)  # 到达率
    ws.cell(row=i, column=start_col + 9, value=0.26)  # 服务时间
    ws.cell(row=i, column=start_col + 10, value=2)  # 服务能力，这里假设它是1到10的整数

# 保存工作簿/
wb.save('resource/option_2_output.xlsx')


